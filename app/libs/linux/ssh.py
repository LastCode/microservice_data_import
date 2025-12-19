from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, TypeVar

import openpyxl
import pandas as pd
import paramiko
from neo4j import GraphDatabase, exceptions as neo4j_exceptions
from openpyxl.styles import Alignment, Font

from app.config.settings import (
    DEFAULT_SETTINGS_PATH,
    Settings,
    SettingsError,
    load_settings,
)

T = TypeVar("T")


@dataclass(frozen=True)
class Neo4JConfig:
    uri: str
    user: str
    password: str
    database: str


class SSHCSVHeaderFetcher:
    domain_map: Mapping[str, str]

    def __init__(self, settings: Settings | str | Path = DEFAULT_SETTINGS_PATH) -> None:
        self.domain_map = {
            "API": "ApiDomain",
            "CALCULATION": "CalculationDomain",
            "CORE": "CoreDomain",
            "CURATED": "CuratedDomain",
            "FEED": "FeedDomain",
            "REPORT": "ReportDomain",
            "REPORT_CATALOG": "ReportCatalogDomain",
            "SCHEDULE": "ScheduleDomain",
        }

        if isinstance(settings, Settings):
            self._settings = settings
        else:
            try:
                self._settings = load_settings(settings)
            except SettingsError as exc:  # pragma: no cover - configuration error path
                raise ValueError(f"Unable to load settings: {exc}") from exc

        self.linux_servers = self._settings.LINUX_SERVERS
        if not self.linux_servers:
            raise ValueError(
                "settings must define LINUX_SERVERS with connection details"
            )

    def _get_neo4j_config(self) -> Neo4JConfig:
        databases = self._settings.DATABASES
        if not databases:
            raise KeyError("settings must define a DATABASES mapping")
        try:
            neo4j_settings = databases["neo4j"]
        except KeyError as exc:  # pragma: no cover - configuration error path
            raise KeyError(
                "DATABASES must include a 'neo4j' configuration block"
            ) from exc

        required_keys = {"NEO4J_URI", "USER", "PASSWORD", "DATABASE"}
        missing = required_keys - neo4j_settings.keys()
        if missing:  # pragma: no cover - configuration error path
            missing_list = ", ".join(sorted(missing))
            raise KeyError(f"Neo4j configuration missing keys: {missing_list}")

        return Neo4JConfig(
            uri=neo4j_settings["NEO4J_URI"],
            user=neo4j_settings["USER"],
            password=neo4j_settings["PASSWORD"],
            database=neo4j_settings["DATABASE"],
        )

    def _with_neo4j_session(self, handler: Callable[[Any], T]) -> T:
        neo4j_cfg = self._get_neo4j_config()
        driver = GraphDatabase.driver(
            neo4j_cfg.uri,
            auth=(neo4j_cfg.user, neo4j_cfg.password),
        )
        try:
            with driver.session(database=neo4j_cfg.database) as session:
                return handler(session)
        except neo4j_exceptions.AuthError as exc:
            raise RuntimeError(
                "Neo4j authentication failed. "
                "Verify DATABASES['neo4j'] credentials in settings.yaml or provide overrides."
            ) from exc
        except neo4j_exceptions.Neo4jError as exc:
            raise RuntimeError(f"Neo4j query failed: {exc}") from exc
        finally:
            driver.close()

    def header_map_to_dataframe(
        self, header_map: Mapping[str, Mapping[str, Iterable[str]]]
    ) -> pd.DataFrame:
        """
        Converts a header_map dictionary to a pandas DataFrame.
        The dictionary should be structured:
        {domain_type: {domain_name: [header_column, ...], ...}, ...}
        """
        rows = []
        for domain_type, domains in header_map.items():
            for domain_name, header_columns in domains.items():
                for col in header_columns:
                    rows.append(
                        {
                            "domain_type": domain_type,
                            "domain_name": domain_name,
                            "header_column": col,
                        }
                    )
        df = pd.DataFrame(rows, columns=["domain_type", "domain_name", "header_column"])
        return df

    def fetch_viper_header(
        self, domain_type: str, domain_name: str
    ) -> Dict[str, Dict[str, List[str]]]:
        """
        Fetch domain businessName headers from Neo4j for a given domain.
        Returns a dictionary:
        {
            domain_type: {
                domain_name: [businessName1, businessName2, ...]
            }
        }
        """
        cypher = """
        MATCH (m:Element)
        WHERE m.domainType = $domain_type AND m.domainName = $domain_name
        RETURN m.physicalName AS physical_name
        """

        def _fetch(session):
            result = session.run(
                cypher,
                domain_type=domain_type,
                domain_name=domain_name,
            )
            cols = [record["physical_name"] for record in result]
            return {domain_type: {domain_name: cols}}

        return self._with_neo4j_session(_fetch)

    def _get_server_config(self, server: str) -> dict:
        try:
            return self.linux_servers[server]
        except KeyError as exc:  # pragma: no cover - configuration error path
            available = ", ".join(sorted(self.linux_servers)) or "<none>"
            raise ValueError(
                f"Unknown server '{server}'. Available servers: {available}"
            ) from exc

    def fetch_file_headers(
        self,
        domain_type: str = "CURATED",
        domain_name: str | None = None,
        server: str | None = None,
    ) -> Dict[str, Dict[str, List[str]]]:
        if domain_name is None:
            raise ValueError("domain_name must be provided when fetching file headers")
        if server is None:
            raise ValueError("server must be provided when fetching file headers")

        server_config = self._get_server_config(server)
        csv_files = server_config.get("csv_files", {}) or {}

        domain_mapping = None
        for key, value in csv_files.items():
            if key.lower() == domain_type.lower():
                domain_mapping = value
                break
        if domain_mapping is None:
            raise ValueError(
                f"No CSV mapping defined for domain type '{domain_type}' on server '{server}'"
            )

        file_info: Mapping[str, Any] | str | None = None
        for key, value in domain_mapping.items():
            if key.strip().lower() == domain_name.strip().lower():
                file_info = value
                break

        if file_info is None:
            raise ValueError(
                f"No file mapping found for domain '{domain_name}' in type '{domain_type}'"
            )

        if isinstance(file_info, str):
            file_info = {"file_path": file_info}

        ssh_user = server_config.get("user")
        ssh_pass = server_config.get("pass")
        if not ssh_user or not ssh_pass:
            raise ValueError(
                f"Missing SSH credentials for server '{server}' in configuration"
            )

        server_host = (
            file_info.get("server_name")
            or file_info.get("server")
            or server_config.get("server_name")
            or server_config.get("host")
            or server_config.get("hostname")
            or server
        )
        file_path = file_info.get("file_path")
        if not file_path:
            raise ValueError(
                f"No file_path defined for domain '{domain_name}' on server '{server}'"
            )
        results = {domain_type: {}}
        ssh: paramiko.SSHClient | None = None
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(server_host, username=ssh_user, password=ssh_pass)
            _stdin, stdout, _stderr = ssh.exec_command(f"head -n 1 {file_path}")
            header_str = stdout.read().decode().strip()
            cols = [col.strip() for col in header_str.split(",") if col.strip()]
            results[domain_type][domain_name] = cols
        except Exception as exc:
            raise RuntimeError(
                f"Failed to fetch headers from {server_host}:{file_path}: {exc}"
            ) from exc
        finally:
            if ssh is not None:
                ssh.close()
        return results

    def autosize(self, ws, min_width=16, pad=2, wrap=True):
        # optional: wrap text for long headers
        if wrap:
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
        # compute max width per column
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells:
                try:
                    val = "" if c.value is None else str(c.value)
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_width, max_len + pad)

    def highlight_summary_sheet(self, ws_sum, is_case_sensitive=False):
        red_font = Font(color="FF0000")
        default_font = Font(color="000000")

        for row in ws_sum.iter_rows(min_row=2, min_col=3, max_col=4):
            left_cell, right_cell = row
            left_val = str(left_cell.value).strip() if left_cell.value else ""
            right_val = str(right_cell.value).strip() if right_cell.value else ""

            # 根據參數決定比較方式
            if not is_case_sensitive:
                left_cmp = left_val.lower()
                right_cmp = right_val.lower()
            else:
                left_cmp = left_val
                right_cmp = right_val

            # Below logic unchanged, just swap to compare the normalized text
            if not left_cmp and right_cmp:
                left_cell.font = default_font
                right_cell.font = red_font
            elif left_cmp and not right_cmp:
                left_cell.font = red_font
                right_cell.font = default_font
            elif left_cmp and right_cmp and left_cmp != right_cmp:
                left_cell.font = red_font
                right_cell.font = red_font
            else:
                left_cell.font = default_font
                right_cell.font = default_font

    def write_summary_to_neo4j(self, summary_df: pd.DataFrame) -> None:
        # 依 domain_type, domain_name 分群存成 lists
        grouped = summary_df.groupby(["domain_type", "domain_name"])
        batch = []
        for (domain_type, domain_name), df_group in grouped:
            viper_list = [
                val if not pd.isna(val) else "" for val in df_group["viper_header"]
            ]
            file_list = [
                val if not pd.isna(val) else "" for val in df_group["file_header"]
            ]
            batch.append(
                {
                    "domain_type": domain_type,
                    "domain_name": domain_name,
                    "viper_headers": viper_list,
                    "file_headers": file_list,
                }
            )

        cypher = """
        UNWIND $batch AS record
        MERGE (s:HeaderMapping {
            domain_type: record.domain_type,
            domain_name: record.domain_name
        })
        SET s.viper_headers = record.viper_headers,
            s.file_headers = record.file_headers
        """

        def _write(session):
            session.run(cypher, batch=batch)

        self._with_neo4j_session(_write)

    def compare_and_report(
        self,
        df1: pd.DataFrame,
        df2: pd.DataFrame,
        domain_type: str | None = None,
        domain_name: str | None = None,
        project_root: str = ".",
        is_case_sensitive: bool = False,
    ) -> pd.DataFrame:
        if domain_type is None:
            raise ValueError("domain_type is required for comparison reports")
        if domain_name is None:
            raise ValueError("domain_name is required for comparison reports")

        now_str = datetime.now().strftime("%Y%m%d")
        output_dir = os.path.join(project_root, "output")
        os.makedirs(output_dir, exist_ok=True)
        report_path = os.path.join(
            output_dir, f"{domain_type}_{domain_name}_{now_str}.xlsx"
        )

        # index: lower-case header to original
        viper_dict = {
            col.lower() if not is_case_sensitive else col: col
            for col in df1["header_column"]
        }
        file_dict = {
            col.lower() if not is_case_sensitive else col: col
            for col in df2["header_column"]
        }

        # 已配對
        matched_keys = set(viper_dict.keys()) & set(file_dict.keys())
        rows = []
        # 先放所有 matched
        for key in sorted(matched_keys):
            rows.append(
                {
                    "domain_type": domain_type,
                    "domain_name": domain_name,
                    "viper_header": viper_dict[key],
                    "file_header": file_dict[key],
                }
            )
        # 再補上 viper only
        for key in sorted(set(viper_dict.keys()) - matched_keys):
            rows.append(
                {
                    "domain_type": domain_type,
                    "domain_name": domain_name,
                    "viper_header": viper_dict[key],
                    "file_header": "",
                }
            )
        # 最後補上 file only
        for key in sorted(set(file_dict.keys()) - matched_keys):
            rows.append(
                {
                    "domain_type": domain_type,
                    "domain_name": domain_name,
                    "viper_header": "",
                    "file_header": file_dict[key],
                }
            )

        summary_df = pd.DataFrame(
            rows, columns=["domain_type", "domain_name", "viper_header", "file_header"]
        )
        self.write_summary_to_neo4j(summary_df)

        with pd.ExcelWriter(report_path) as writer:
            df1.to_excel(writer, sheet_name="viper_headers", index=False)
            df2.to_excel(writer, sheet_name="file_headers", index=False)
            summary_df.to_excel(writer, sheet_name="summary", index=False)
        wb = openpyxl.load_workbook(report_path)
        ws_sum = wb["summary"]

        # Highlight missing cells and adjust column widths in summary
        self.highlight_summary_sheet(ws_sum)
        self.autosize(ws_sum)
        self.autosize(wb["viper_headers"])
        self.autosize(wb["file_headers"])

        wb.save(report_path)
        print(f"Comparison report saved to {report_path}")

        return summary_df, report_path
